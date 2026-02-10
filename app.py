from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
import json
import os
import csv
import re
from io import BytesIO
from difflib import SequenceMatcher, get_close_matches
from werkzeug.utils import secure_filename
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from auth import auth, get_user_by_id
from decorators import login_required, admin_required
from models import FAQ, KnowledgeBaseArticle, SupportTicket, StatusUpdate, SupportJobCard, EscalatedJobCard

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'

# Register auth blueprint
app.register_blueprint(auth)

# Configure upload folder
UPLOAD_FOLDER = os.path.join('static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB max file size

# Create uploads folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

@app.context_processor
def inject_cart_count():
    """Make cart item count available in all templates for the navbar badge"""
    cart = load_cart() if os.path.exists(CART_FILE) else []
    return dict(global_cart_count=sum(item.get("quantity", 0) for item in cart))

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Job Cards Storage
DATA_FILE = "jobs_data.json"

def load_jobs():
    """Load jobs from JSON file"""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return [
        {"job_id": "JC1234", "status": "Open",   "remarks": "Initial review", "assigned_to": "Alice"},
        {"job_id": "JC5678", "status": "Closed", "remarks": "Completed",      "assigned_to": "Bob"},
    ]

def save_jobs(jobs):
    """Save jobs to JSON file"""
    with open(DATA_FILE, "w") as f:
        json.dump(jobs, f, indent=2)

jobs = load_jobs()

# Orders Storage
ORDERS_FILE = "orders_data.json"
CART_FILE = "cart_data.json"

def load_orders():
    """Load orders from JSON file"""
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_orders(orders):
    """Save orders to JSON file"""
    with open(ORDERS_FILE, "w") as f:
        json.dump(orders, f, indent=2)

def load_cart():
    """Load shopping cart from JSON file"""
    if os.path.exists(CART_FILE):
        with open(CART_FILE, "r") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_cart(cart):
    """Save shopping cart to JSON file"""
    with open(CART_FILE, "w") as f:
        json.dump(cart, f, indent=2)

# Rewards Storage
REWARDS_FILE = "rewards_data.json"

def load_rewards():
    """
    Stored format (list of dicts):
    [
      {"customer_id": "S1234567A", "name": "Alice Tan", "purchases": 2, "redeemed": 0}
    ]
    """
    if os.path.exists(REWARDS_FILE):
        with open(REWARDS_FILE, "r") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_rewards(customers):
    with open(REWARDS_FILE, "w") as f:
        json.dump(customers, f, indent=2)

def compute_reward_balance(customer):
    """
    1 discount per 3 purchases.
    balance = floor(purchases/3) - redeemed
    """
    earned = customer.get("purchases", 0) // 3
    redeemed = customer.get("redeemed", 0)
    return max(0, earned - redeemed)

# Catalogue Storage
CATALOGUE_FILE = "catalogue_data.json"

def load_catalogue():
    """Load catalogue items from JSON file"""
    if os.path.exists(CATALOGUE_FILE):
        with open(CATALOGUE_FILE, "r") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_catalogue(catalogue):
    """Save catalogue items to JSON file"""
    with open(CATALOGUE_FILE, "w") as f:
        json.dump(catalogue, f, indent=2)

# ==================== SUPPORT FEATURES STORAGE ====================

# Support Tickets Storage
TICKETS_FILE = "support_tickets_data.json"

def load_support_tickets():
    """Load support tickets from JSON file"""
    if os.path.exists(TICKETS_FILE):
        with open(TICKETS_FILE, "r") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def save_support_tickets(tickets):
    """Save support tickets to JSON file"""
    with open(TICKETS_FILE, "w") as f:
        json.dump(tickets, f, indent=2)

# FAQs stored in memory
faqs_data = [
    FAQ(1, "How do I reset my password?", "Visit the login page and click 'Forgot Password'. Enter your email and follow the instructions sent to your inbox.", "account").to_dict(),
    FAQ(2, "What are your support hours?", "We provide 24/7 support via email and ticket system. AI Support is available 9 AM - 5 PM EST.", "general").to_dict(),
    FAQ(3, "How long does billing take to process?", "Invoices are processed within 1-2 business days. You'll receive a confirmation email.", "billing").to_dict(),
    FAQ(4, "Can I cancel my subscription?", "Yes, you can cancel anytime from your account settings. Your access continues until the end of the billing period.", "billing").to_dict(),
    FAQ(5, "Is my data secure?", "We use 256-bit SSL encryption and comply with GDPR and industry security standards.", "account").to_dict(),
]

# Knowledge Base articles stored in memory
kb_data = [
    KnowledgeBaseArticle(1, "Getting Started Guide", "Step 1: Create an account\nStep 2: Verify your email\nStep 3: Set up your profile\nStep 4: Explore the dashboard", "guide", "beginner").to_dict(),
    KnowledgeBaseArticle(2, "Troubleshooting Login Issues", "If you can't login:\n1. Clear your browser cache\n2. Try incognito mode\n3. Check if caps lock is on\n4. Reset password if needed", "troubleshooting", "beginner").to_dict(),
    KnowledgeBaseArticle(3, "Advanced Settings Configuration", "Configure API keys, webhooks, and integrations in your settings panel. Requires authentication.", "guide", "advanced").to_dict(),
]

# Status updates stored in memory
status_data = [
    StatusUpdate(1, "All Systems Operational", "All services running normally.", "operational", "low").to_dict(),
]

def load_faqs():
    """In-Memory: Load FAQs"""
    return faqs_data

def load_kb():
    """In-Memory: Load knowledge base articles"""
    return kb_data

def load_status():
    """In-Memory: Load status updates"""
    return status_data

@app.route("/", methods=["GET", "POST"])
@login_required
def home():
    global jobs
    jobs = load_jobs()

    search_query = request.args.get("search", "").strip().lower()
    is_admin = session.get('role') == 'admin'

    if request.method == "POST":
        # Only admins can create job cards
        if not is_admin:
            flash('Only admins can create job cards', 'danger')
            return redirect(url_for("home"))
        
        job_id = request.form.get("job_id", "").strip()
        status = request.form.get("status", "").strip()
        remarks = request.form.get("remarks", "").strip()
        assigned_to = request.form.get("assigned_to", "").strip()

        if job_id and status:
            # Prevent duplicate Job ID
            if any(j.get("job_id") == job_id for j in jobs):
                return redirect(url_for("home")) 
            jobs.append({
                "job_id": job_id,
                "status": status,
                "remarks": remarks,
                "assigned_to": assigned_to,
            })
            save_jobs(jobs)

        return redirect(url_for("home"))

    filtered_jobs = jobs
    if search_query:
        filtered_jobs = [
            job for job in jobs
            if search_query in job.get("job_id", "").lower()
            or search_query in job.get("status", "").lower()
            or search_query in job.get("remarks", "").lower()
            or search_query in job.get("assigned_to", "").lower()
        ]

    return render_template("index.html", jobs=filtered_jobs, search_query=search_query, is_admin=is_admin)

@app.route("/update/<job_id>", methods=["GET", "POST"])
@admin_required
def update(job_id):
    global jobs
    jobs = load_jobs()

    job = next((j for j in jobs if j["job_id"] == job_id), None)
    if not job:
        return redirect(url_for("home"))

    if request.method == "POST":
        job["status"] = request.form.get("status", "").strip()
        job["remarks"] = request.form.get("remarks", "").strip()
        job["assigned_to"] = request.form.get("assigned_to", "").strip()
        save_jobs(jobs)
        return redirect(url_for("home"))

    return render_template("update.html", job=job)

@app.route("/delete/<job_id>", methods=["POST"])
@admin_required
def delete(job_id):
    global jobs
    jobs = load_jobs()
    jobs = [j for j in jobs if j["job_id"] != job_id]
    save_jobs(jobs)
    return redirect(url_for("home"))

@app.route("/rewards", methods=["GET", "POST"])
@admin_required
def rewards():
    customers = load_rewards()
    search_query = request.args.get("search", "").strip().lower()

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        # Create/Update customer
        if action == "create_customer":
            customer_id = request.form.get("customer_id", "").strip()
            name = request.form.get("name", "").strip()

            if customer_id:
                existing = next((c for c in customers if c["customer_id"] == customer_id), None)
                if not existing:
                    customers.append({
                        "customer_id": customer_id,
                        "name": name,
                        "purchases": 0,
                        "redeemed": 0
                    })
                else:
                    # allow updating name if provided
                    if name:
                        existing["name"] = name

                save_rewards(customers)

            return redirect(url_for("rewards"))

        # Add purchase to a customer
        if action == "add_purchase":
            customer_id = request.form.get("customer_id", "").strip()
            amount = request.form.get("amount", "1").strip()

            try:
                amount_int = int(amount)
            except ValueError:
                amount_int = 1

            if amount_int < 1:
                amount_int = 1

            cust = next((c for c in customers if c["customer_id"] == customer_id), None)
            if cust:
                cust["purchases"] = int(cust.get("purchases", 0)) + amount_int
                save_rewards(customers)

            return redirect(url_for("rewards"))

    # filter/search
    filtered = customers
    if search_query:
        filtered = [
            c for c in customers
            if search_query in c.get("customer_id", "").lower()
            or search_query in c.get("name", "").lower()
        ]

    # add computed fields for display
    display_customers = []
    for c in filtered:
        c2 = dict(c)
        c2["reward_balance"] = compute_reward_balance(c)
        c2["next_reward_at"] = ((c.get("purchases", 0) // 3) + 1) * 3  # next threshold
        display_customers.append(c2)

    # sort by most purchases desc
    display_customers.sort(key=lambda x: x.get("purchases", 0), reverse=True)

    return render_template(
        "rewards.html",
        customers=display_customers,
        search_query=search_query
    )

@app.route("/rewards/redeem/<customer_id>", methods=["POST"])
@admin_required
def redeem_reward(customer_id):
    customers = load_rewards()
    cust = next((c for c in customers if c["customer_id"] == customer_id), None)
    if cust:
        if compute_reward_balance(cust) > 0:
            cust["redeemed"] = int(cust.get("redeemed", 0)) + 1
            save_rewards(customers)
    return redirect(url_for("rewards"))

# ==================== CATALOGUE ROUTES ====================

@app.route("/catalogue", methods=["GET", "POST"])
@login_required
def catalogue():
    """Display and manage parts catalogue with images"""
    catalogue_items = load_catalogue()
    search_query = request.args.get("search", "").strip().lower()
    category_filter = request.args.get("category", "").strip()
    
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        
        # Add new part
        if action == "add_part":
            part_id = request.form.get("part_id", "").strip()
            name = request.form.get("name", "").strip()
            category = request.form.get("category", "").strip()
            price = request.form.get("price", "0").strip()
            stock = request.form.get("stock", "0").strip()
            description = request.form.get("description", "").strip()
            
            # Handle image upload
            image_path = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename and allowed_file(file.filename):
                    # Use part_id if available, otherwise use timestamp
                    prefix = part_id if part_id else str(int(__import__('time').time()))
                    filename = secure_filename(f"{prefix}_{file.filename}")
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    try:
                        file.save(filepath)
                        image_path = f"/static/uploads/{filename}"
                    except Exception as e:
                        print(f"Error saving file: {e}")
                        image_path = None
            
            # Validate required fields
            if part_id and name and category:
                try:
                    price_float = float(price)
                    stock_int = int(stock)
                except ValueError:
                    price_float = 0.0
                    stock_int = 0
                
                new_part = {
                    "part_id": part_id,
                    "name": name,
                    "category": category,
                    "price": price_float,
                    "stock": stock_int,
                    "description": description,
                    "image": image_path
                }
                
                # Check if part already exists
                existing = next((p for p in catalogue_items if p["part_id"] == part_id), None)
                if existing:
                    # Update existing part
                    existing.update(new_part)
                else:
                    # Add new part
                    catalogue_items.append(new_part)
                
                save_catalogue(catalogue_items)
        
        # Update stock
        elif action == "update_stock":
            part_id = request.form.get("part_id", "").strip()
            stock = request.form.get("stock", "0").strip()
            
            try:
                stock_int = int(stock)
            except ValueError:
                stock_int = 0
            
            part = next((p for p in catalogue_items if p["part_id"] == part_id), None)
            if part:
                part["stock"] = stock_int
                save_catalogue(catalogue_items)
        
        # Delete part
        elif action == "delete_part":
            part_id = request.form.get("part_id", "").strip()
            catalogue_items = [p for p in catalogue_items if p["part_id"] != part_id]
            save_catalogue(catalogue_items)
        
        # Add to cart from catalogue
        elif action == "add_to_cart_from_catalogue":
            part_id = request.form.get("part_id", "").strip()
            quantity = request.form.get("quantity", "1").strip()
            
            part = next((p for p in catalogue_items if p["part_id"] == part_id), None)
            if part:
                try:
                    quantity_int = int(quantity)
                except ValueError:
                    quantity_int = 1
                
                cart = load_cart()
                existing = next((item for item in cart if item["item_name"] == part["name"]), None)
                if existing:
                    existing["quantity"] += quantity_int
                else:
                    cart.append({
                        "item_name": part["name"],
                        "price": part["price"],
                        "quantity": quantity_int
                    })
                save_cart(cart)
                flash(f'Added {quantity_int}x {part["name"]} to cart!', 'success')
        
        return redirect(url_for("catalogue", search=search_query, category=category_filter))
    
    # Get unique categories
    categories = sorted(set(p.get("category", "Other") for p in catalogue_items if p.get("category")))
    
    # Filter catalogue
    filtered_catalogue = catalogue_items
    
    if category_filter:
        filtered_catalogue = [p for p in filtered_catalogue if p.get("category") == category_filter]
    
    # ========== AI SMART SEARCH with Fuzzy Matching ==========
    search_suggestions = []
    fuzzy_matches = []
    
    if search_query:
        # Exact substring match first
        exact_matches = [
            p for p in filtered_catalogue
            if search_query in p.get("part_id", "").lower()
            or search_query in p.get("name", "").lower()
            or search_query in p.get("category", "").lower()
            or search_query in p.get("description", "").lower()
        ]
        
        if exact_matches:
            filtered_catalogue = exact_matches
        else:
            # Fuzzy matching - find similar items when exact search fails
            all_names = [p.get("name", "") for p in filtered_catalogue]
            all_categories = list(set(p.get("category", "") for p in filtered_catalogue))
            all_searchable = all_names + all_categories
            
            # Get close matches for the search query
            close_names = get_close_matches(search_query, [n.lower() for n in all_names], n=5, cutoff=0.4)
            close_cats = get_close_matches(search_query, [c.lower() for c in all_categories], n=3, cutoff=0.4)
            
            # Also do character-level similarity
            for p in filtered_catalogue:
                name_ratio = SequenceMatcher(None, search_query, p.get("name", "").lower()).ratio()
                cat_ratio = SequenceMatcher(None, search_query, p.get("category", "").lower()).ratio()
                desc_ratio = SequenceMatcher(None, search_query, p.get("description", "").lower()).ratio()
                best_ratio = max(name_ratio, cat_ratio, desc_ratio)
                if best_ratio > 0.35:
                    fuzzy_matches.append((p, best_ratio))
            
            # Sort fuzzy matches by relevance
            fuzzy_matches.sort(key=lambda x: x[1], reverse=True)
            
            if fuzzy_matches:
                filtered_catalogue = [m[0] for m in fuzzy_matches]
                # Generate "Did you mean?" suggestions
                search_suggestions = list(set(close_names + close_cats))[:5]
            else:
                filtered_catalogue = []
                search_suggestions = list(set(close_names + close_cats))[:5]
    
    # ========== SORT OPTIONS ==========
    sort_by = request.args.get("sort", "").strip()
    if sort_by == "price_low":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("price", 0))
    elif sort_by == "price_high":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("price", 0), reverse=True)
    elif sort_by == "name_az":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("name", "").lower())
    elif sort_by == "name_za":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("name", "").lower(), reverse=True)
    elif sort_by == "stock_low":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("stock", 0))
    elif sort_by == "stock_high":
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: p.get("stock", 0), reverse=True)
    else:
        # Default: Sort by part_id (natural sort)
        def extract_number(part_id):
            match = re.search(r'\d+', part_id)
            return int(match.group()) if match else 0
        filtered_catalogue = sorted(filtered_catalogue, key=lambda p: extract_number(p.get("part_id", "")))
    
    # ========== ANALYTICS DATA ==========
    total_parts = len(catalogue_items)
    total_stock = sum(p.get("stock", 0) for p in catalogue_items)
    total_value = sum(p.get("price", 0) * p.get("stock", 0) for p in catalogue_items)
    avg_price = sum(p.get("price", 0) for p in catalogue_items) / max(total_parts, 1)
    low_stock_parts = [p for p in catalogue_items if 0 < p.get("stock", 0) <= 5]
    out_of_stock_parts = [p for p in catalogue_items if p.get("stock", 0) == 0]
    
    # Category distribution for chart
    category_stats = {}
    for p in catalogue_items:
        cat = p.get("category", "Other")
        if cat not in category_stats:
            category_stats[cat] = {"count": 0, "total_stock": 0, "total_value": 0}
        category_stats[cat]["count"] += 1
        category_stats[cat]["total_stock"] += p.get("stock", 0)
        category_stats[cat]["total_value"] += p.get("price", 0) * p.get("stock", 0)
    
    # ========== AI RECOMMENDATIONS ==========
    # Build category relationship map for "You might also need" suggestions
    recommendations = []
    cart = load_cart()
    cart_item_names = [item["item_name"] for item in cart]
    
    # Get categories of items in cart
    cart_categories = set()
    for item in cart:
        for p in catalogue_items:
            if p["name"] == item["item_name"]:
                cart_categories.add(p.get("category", ""))
    
    # Recommend items from same categories not already in cart
    if cart_categories:
        for p in catalogue_items:
            if p["name"] not in cart_item_names and p.get("category") in cart_categories and p.get("stock", 0) > 0:
                recommendations.append(p)
    
    # If no cart-based recommendations, suggest popular/high-stock items
    if not recommendations:
        recommendations = sorted(
            [p for p in catalogue_items if p.get("stock", 0) > 0],
            key=lambda p: p.get("stock", 0),
            reverse=True
        )[:4]
    else:
        recommendations = recommendations[:4]
    
    # Cart count for badge
    cart_item_count = sum(item.get("quantity", 0) for item in cart)
    
    return render_template(
        "catalogue.html",
        catalogue=filtered_catalogue,
        categories=categories,
        search_query=search_query,
        category_filter=category_filter,
        sort_by=sort_by,
        search_suggestions=search_suggestions,
        total_parts=total_parts,
        total_stock=total_stock,
        total_value=total_value,
        avg_price=avg_price,
        low_stock_parts=low_stock_parts,
        out_of_stock_parts=out_of_stock_parts,
        category_stats=category_stats,
        recommendations=recommendations,
        cart_item_count=cart_item_count
    )

@app.route("/catalogue/edit/<part_id>", methods=["GET", "POST"])
@admin_required
def edit_catalogue_part(part_id):
    """Edit a catalogue part"""
    catalogue_items = load_catalogue()
    part = next((p for p in catalogue_items if p["part_id"] == part_id), None)
    
    if not part:
        return redirect(url_for("catalogue"))
    
    if request.method == "POST":
        # Update part details
        part["name"] = request.form.get("name", "").strip()
        part["category"] = request.form.get("category", "").strip()
        part["description"] = request.form.get("description", "").strip()
        
        try:
            part["price"] = float(request.form.get("price", "0"))
        except ValueError:
            part["price"] = 0.0
        
        try:
            part["stock"] = int(request.form.get("stock", "0"))
        except ValueError:
            part["stock"] = 0
        
        # Handle image upload
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{part_id}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                try:
                    file.save(filepath)
                    part["image"] = f"/static/uploads/{filename}"
                except Exception as e:
                    print(f"Error saving file: {e}")
        
        save_catalogue(catalogue_items)
        return redirect(url_for("catalogue"))
    
    return render_template("edit_part.html", part=part)

@app.route("/catalogue/export", methods=["GET"])
@admin_required
def export_catalogue():
    """Export catalogue to CSV"""
    from datetime import datetime
    catalogue_items = load_catalogue()
    
    # Create CSV in memory using BytesIO for binary mode
    output = BytesIO()
    text_stream = output.write
    
    # Manually write CSV content as bytes
    lines = []
    lines.append("Part ID,Name,Category,Price,Stock,Description,Image\n")
    
    for part in catalogue_items:
        # Escape quotes and wrap fields with quotes
        row = [
            f'"{part.get("part_id", "").replace(chr(34), chr(34)+chr(34))}"',
            f'"{part.get("name", "").replace(chr(34), chr(34)+chr(34))}"',
            f'"{part.get("category", "").replace(chr(34), chr(34)+chr(34))}"',
            str(part.get("price", "")),
            str(part.get("stock", "")),
            f'"{part.get("description", "").replace(chr(34), chr(34)+chr(34))}"',
            f'"{part.get("image", "").replace(chr(34), chr(34)+chr(34))}"'
        ]
        lines.append(",".join(row) + "\n")
    
    # Write as bytes
    csv_content = "".join(lines)
    output.write(csv_content.encode('utf-8'))
    output.seek(0)
    
    filename = f"catalogue_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv"
    )

@app.route("/catalogue/import", methods=["POST"])
@admin_required
def import_catalogue():
    """Import parts from CSV or Excel file"""
    try:
        if 'file' not in request.files:
            return redirect(url_for('catalogue'))
        
        file = request.files['file']
        if not file or file.filename == '':
            return redirect(url_for('catalogue'))
        
        filename = file.filename.lower()
        catalogue_items = load_catalogue()
        rows_processed = 0
        rows_added = 0
        rows_updated = 0
        
        # Handle Excel files
        if filename.endswith(('.xlsx', '.xls')):
            if not HAS_OPENPYXL:
                return redirect(url_for('catalogue'))
            
            try:
                workbook = load_workbook(file)
                worksheet = workbook.active
                
                for idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                    if idx == 1:  # Skip header
                        continue
                    
                    if not row or not row[0]:  # Skip empty rows
                        continue
                    
                    rows_processed += 1
                    part_id = str(row[0]).strip()
                    
                    # Check if part exists
                    existing = next((p for p in catalogue_items if p["part_id"] == part_id), None)
                    
                    # Parse values
                    try:
                        price = float(row[3]) if row[3] else 0.0
                    except (ValueError, TypeError):
                        price = 0.0
                    
                    try:
                        stock = int(row[4]) if row[4] else 0
                    except (ValueError, TypeError):
                        stock = 0
                    
                    part_data = {
                        "part_id": part_id,
                        "name": str(row[1]).strip() if row[1] else "",
                        "category": str(row[2]).strip() if row[2] else "",
                        "price": price,
                        "stock": stock,
                        "description": str(row[5]).strip() if len(row) > 5 and row[5] else "",
                        "image": str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    }
                    
                    if existing:
                        existing.update(part_data)
                        rows_updated += 1
                    else:
                        catalogue_items.append(part_data)
                        rows_added += 1
                
                workbook.close()
            except Exception as e:
                print(f"Error reading Excel: {e}")
                return redirect(url_for('catalogue'))
        
        # Handle CSV files
        elif filename.endswith('.csv'):
            try:
                file.seek(0)
                stream = file.read().decode('utf-8')
                reader = csv.reader(stream.splitlines())
                
                for idx, row in enumerate(reader):
                    if idx == 0:  # Skip header
                        continue
                    
                    if not row or not row[0].strip():
                        continue
                    
                    rows_processed += 1
                    part_id = row[0].strip()
                    
                    # Check if part exists
                    existing = next((p for p in catalogue_items if p["part_id"] == part_id), None)
                    
                    # Parse values
                    try:
                        price = float(row[3]) if len(row) > 3 and row[3] else 0.0
                    except (ValueError, TypeError):
                        price = 0.0
                    
                    try:
                        stock = int(row[4]) if len(row) > 4 and row[4] else 0
                    except (ValueError, TypeError):
                        stock = 0
                    
                    part_data = {
                        "part_id": part_id,
                        "name": row[1].strip() if len(row) > 1 else "",
                        "category": row[2].strip() if len(row) > 2 else "",
                        "price": price,
                        "stock": stock,
                        "description": row[5].strip() if len(row) > 5 else "",
                        "image": row[6].strip() if len(row) > 6 else ""
                    }
                    
                    if existing:
                        existing.update(part_data)
                        rows_updated += 1
                    else:
                        catalogue_items.append(part_data)
                        rows_added += 1
            except Exception as e:
                print(f"Error reading CSV: {e}")
                flash(f"Error reading CSV file: {str(e)}", "error")
                return redirect(url_for('catalogue'))
        
        # Save updated catalogue
        if rows_processed > 0:
            save_catalogue(catalogue_items)
            flash(f"Import successful! Added {rows_added} parts, Updated {rows_updated} parts", "success")
        else:
            flash("No valid data found in file", "warning")
        
        return redirect(url_for('catalogue'))
    except Exception as e:
        print(f"Error importing catalogue: {e}")
        flash(f"Error importing catalogue: {str(e)}", "error")
        return redirect(url_for('catalogue'))

# ==================== ORDERS & CART ROUTES ====================

@app.route("/orders", methods=["GET", "POST"])
@login_required
def orders():
    """Display orders and shopping cart"""
    orders_list = load_orders()
    cart = load_cart()
    search_query = request.args.get("search", "").strip().lower()

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        # Add item to cart
        if action == "add_to_cart":
            item_name = request.form.get("item_name", "").strip()
            price = request.form.get("price", "0").strip()
            quantity = request.form.get("quantity", "1").strip()

            try:
                price_float = float(price)
                quantity_int = int(quantity)
            except ValueError:
                price_float = 0.0
                quantity_int = 1

            if item_name and price_float > 0:
                # Check if item already in cart
                existing = next((item for item in cart if item["item_name"] == item_name), None)
                if existing:
                    existing["quantity"] += quantity_int
                else:
                    cart.append({
                        "item_name": item_name,
                        "price": price_float,
                        "quantity": quantity_int
                    })
                save_cart(cart)

            return redirect(url_for("orders"))

        # Remove item from cart
        if action == "remove_from_cart":
            item_name = request.form.get("item_name", "").strip()
            cart = [item for item in cart if item["item_name"] != item_name]
            save_cart(cart)
            return redirect(url_for("orders"))

        # Update cart item quantity
        if action == "update_cart_quantity":
            item_name = request.form.get("item_name", "").strip()
            quantity = request.form.get("quantity", "1").strip()

            try:
                quantity_int = int(quantity)
            except ValueError:
                quantity_int = 1

            if quantity_int < 1:
                quantity_int = 1

            for item in cart:
                if item["item_name"] == item_name:
                    item["quantity"] = quantity_int
                    break

            save_cart(cart)
            return redirect(url_for("orders"))

        # Checkout - convert cart to order
        if action == "checkout":
            if cart:
                import datetime
                order_id = f"ORD{len(orders_list) + 1001}"
                
                total = sum(item["price"] * item["quantity"] for item in cart)
                
                # Get customer details
                customer_name = request.form.get("customer_name", "Guest").strip()
                customer_phone = request.form.get("customer_phone", "").strip()
                customer_email = request.form.get("customer_email", "").strip()
                order_notes = request.form.get("order_notes", "").strip()
                
                # Delivery option
                delivery_option = request.form.get("delivery_option", "pickup").strip()
                delivery_address = ""
                delivery_fee = 0.0
                if delivery_option == "delivery":
                    delivery_address = request.form.get("delivery_address", "").strip()
                    delivery_fee = 5.00
                    total += delivery_fee
                
                new_order = {
                    "order_id": order_id,
                    "customer_name": customer_name,
                    "customer_phone": customer_phone,
                    "customer_email": customer_email,
                    "order_notes": order_notes,
                    "items": cart.copy(),
                    "total": round(total, 2),
                    "status": "Pending",
                    "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "delivery_option": delivery_option,
                    "delivery_address": delivery_address,
                    "delivery_fee": delivery_fee,
                    "delivery_status": "Preparing" if delivery_option == "delivery" else "",
                    "placed_by": session.get("username", "Unknown")
                }
                
                orders_list.append(new_order)
                save_orders(orders_list)
                
                # Clear cart after checkout
                save_cart([])
                cart = []

            return redirect(url_for("orders"))

        # Update order status
        if action == "update_order_status":
            order_id = request.form.get("order_id", "").strip()
            new_status = request.form.get("status", "").strip()

            order = next((o for o in orders_list if o["order_id"] == order_id), None)
            if order and new_status:
                order["status"] = new_status
                save_orders(orders_list)

            return redirect(url_for("orders"))

    # Calculate cart totals
    cart_total = sum(item["price"] * item["quantity"] for item in cart)
    cart_item_count = sum(item["quantity"] for item in cart)

    # Filter orders by search
    filtered_orders = orders_list
    if search_query:
        filtered_orders = [
            o for o in orders_list
            if search_query in o.get("order_id", "").lower()
            or search_query in o.get("customer_name", "").lower()
            or search_query in o.get("status", "").lower()
        ]

    # Sort orders by most recent
    filtered_orders = sorted(filtered_orders, key=lambda x: x.get("date", ""), reverse=True)

    return render_template(
        "orders.html",
        orders=filtered_orders,
        cart=cart,
        cart_total=round(cart_total, 2),
        cart_item_count=cart_item_count,
        search_query=search_query
    )

@app.route("/orders/delete/<order_id>", methods=["POST"])
@login_required
def delete_order(order_id):
    """Delete an order"""
    orders_list = load_orders()
    orders_list = [o for o in orders_list if o["order_id"] != order_id]
    save_orders(orders_list)
    return redirect(url_for("orders"))

@app.route("/orders/edit/<order_id>", methods=["GET", "POST"])
@login_required
def edit_order(order_id):
    """Edit an existing order"""
    orders_list = load_orders()
    order = next((o for o in orders_list if o["order_id"] == order_id), None)
    
    if not order:
        return redirect(url_for("orders"))
    
    if request.method == "POST":
        # Update order details
        order["customer_name"] = request.form.get("customer_name", "").strip()
        order["customer_phone"] = request.form.get("customer_phone", "").strip()
        order["customer_email"] = request.form.get("customer_email", "").strip()
        order["order_notes"] = request.form.get("order_notes", "").strip()
        order["status"] = request.form.get("status", "Pending").strip()
        
        save_orders(orders_list)
        return redirect(url_for("orders"))
    
    return render_template("edit_order.html", order=order)

# ==================== SUPPORT FEATURES ROUTES ====================

@app.route("/faq")
def faq():
    """FAQ section"""
    faqs = load_faqs()
    search_query = request.args.get("q", "").lower()
    category_filter = request.args.get("category", "")
    
    # Filtering logic
    if search_query:
        faqs = [f for f in faqs if search_query in f.get("question", "").lower() or search_query in f.get("answer", "").lower()]
    if category_filter:
        faqs = [f for f in faqs if f.get("category") == category_filter]
    
    categories = set(f.get("category") for f in load_faqs())
    return render_template("faq.html", faqs=faqs, search_query=search_query, categories=categories, selected_category=category_filter)

@app.route("/knowledge-base")
def knowledge_base():
    """Knowledge base section"""
    articles = load_kb()
    search_query = request.args.get("q", "").lower()
    category_filter = request.args.get("category", "")
    difficulty_filter = request.args.get("difficulty", "")
    
    # Filtering logic
    if search_query:
        articles = [a for a in articles if search_query in a.get("title", "").lower() or search_query in a.get("content", "").lower()]
    if category_filter:
        articles = [a for a in articles if a.get("category") == category_filter]
    if difficulty_filter:
        articles = [a for a in articles if a.get("difficulty") == difficulty_filter]
    
    categories = set(a.get("category") for a in load_kb())
    difficulties = set(a.get("difficulty") for a in load_kb())
    return render_template("knowledge_base.html", articles=articles, search_query=search_query, categories=categories, difficulties=difficulties, selected_category=category_filter, selected_difficulty=difficulty_filter)

@app.route("/article/<int:article_id>")
def article_detail(article_id):
    """View single article"""
    articles = load_kb()
    article = next((a for a in articles if a.get("article_id") == article_id), None)
    if not article:
        flash("Article not found!", "danger")
        return redirect(url_for("knowledge_base"))
    return render_template("article_detail.html", article=article)

@app.route("/tickets", methods=["GET", "POST"])
def tickets():
    """Support tickets page - Users can view and submit tickets only"""
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        issue_type = request.form.get("issue_type", "").strip()
        description = request.form.get("description", "").strip()
        
        # Validation
        if not name or not email or not issue_type or not description:
            flash("Name, Email, Issue Type, and Description are required!", "danger")
            return redirect(url_for("tickets"))
        
        ticket = SupportTicket(name, email, issue_type, description, phone)
        tickets_list = load_support_tickets()
        tickets_list.append(ticket.to_dict())
        save_support_tickets(tickets_list)
        flash(f"Support ticket {ticket.ticket_id} created successfully! We'll respond within 24 hours.", "success")
        return redirect(url_for("tickets"))
    
    all_tickets = load_support_tickets()
    return render_template("tickets.html", tickets=all_tickets)

@app.route("/status")
def status():
    """Service status page"""
    updates = load_status()
    return render_template("status.html", updates=updates)

@app.route("/contact")
def contact():
    """Contact page with multiple channels"""
    chat_url = app.config.get('CHATGPT_CHAT_URL') or os.environ.get('CHATGPT_CHAT_URL') or 'https://chat.openai.com'
    return render_template("contact.html", chat_url=chat_url)


@app.route("/ai-support")
def ai_support():
    """Redirect to configured ChatGPT/AI support URL (open in new tab)."""
    url = app.config.get('CHATGPT_CHAT_URL') or os.environ.get('CHATGPT_CHAT_URL') or 'https://chat.openai.com'
    return redirect(url)


@app.route("/self-service")
@login_required
def self_service():
    """Self-service tools page"""
    return render_template("self_service.html")


@app.route("/search")
@login_required
def search():
    """Global search across FAQs, KB, and other content"""
    query = request.args.get("q", "").lower()
    faqs = load_faqs()
    articles = load_kb()

    results = {
        "faqs": [f for f in faqs if query in f.get("question", "").lower() or query in f.get("answer", "").lower()] if query else [],
        "articles": [a for a in articles if query in a.get("title", "").lower() or query in a.get("content", "").lower()] if query else []
    }

    return render_template("search_results.html", query=query, results=results)


@app.route("/tickets/edit/<ticket_id>", methods=["GET", "POST"])
@login_required
def edit_ticket(ticket_id):
    """Edit an existing support ticket"""
    tickets_list = load_support_tickets()
    ticket = next((t for t in tickets_list if t.get("ticket_id") == ticket_id), None)
    if not ticket:
        flash("Ticket not found!", "danger")
        return redirect(url_for("tickets"))

    if request.method == "POST":
        ticket["name"] = request.form.get("name", "").strip()
        ticket["email"] = request.form.get("email", "").strip()
        ticket["phone"] = request.form.get("phone", "").strip()
        ticket["issue_type"] = request.form.get("issue_type", "").strip()
        ticket["description"] = request.form.get("description", "").strip()

        save_support_tickets(tickets_list)
        flash(f"Support ticket {ticket_id} updated successfully!", "success")
        return redirect(url_for("tickets"))

    return render_template("edit_ticket.html", ticket=ticket)


@app.errorhandler(404)
def page_not_found(e):
    """Custom 404 error page"""
    return render_template("404.html"), 404


# ==================== DELIVERY MANAGEMENT ====================

@app.route("/deliveries", methods=["GET", "POST"])
@login_required
def deliveries():
    """Delivery management page"""
    orders_list = load_orders()
    
    # Filter only delivery orders
    delivery_orders = [o for o in orders_list if o.get("delivery_option") == "delivery"]
    
    if request.method == "POST" and session.get("role") == "admin":
        action = request.form.get("action", "").strip()
        
        if action == "update_delivery_status":
            order_id = request.form.get("order_id", "").strip()
            new_status = request.form.get("delivery_status", "").strip()
            
            order = next((o for o in orders_list if o["order_id"] == order_id), None)
            if order and new_status:
                order["delivery_status"] = new_status
                save_orders(orders_list)
            
            return redirect(url_for("deliveries"))
    
    # For regular users, only show their own orders
    if session.get("role") != "admin":
        current_user = session.get("username", "")
        delivery_orders = [o for o in delivery_orders if o.get("placed_by") == current_user]
    
    # Sort by most recent
    delivery_orders = sorted(delivery_orders, key=lambda x: x.get("date", ""), reverse=True)
    
    return render_template(
        "deliveries.html",
        delivery_orders=delivery_orders,
        is_admin=session.get("role") == "admin"
    )


if __name__ == "__main__":
    app.run(debug=True)
